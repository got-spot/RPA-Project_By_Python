import openpyxl 
import io
from openpyxl.utils import get_column_letter


def appendSheet(ws_dest, target_row, target_col, ws_src, row_offset = 1, col_offset = 1, no_col = 3):
    # calculate total number of rows and  
    # columns in source excel file 
    mr = ws_src.max_row
    mc = no_col
    #mc = ws_src.max_column - 3
    

    # copying the cell values from source  
    # excel file to destination excel file 
    for i in range (0, mr): 
        for j in range (0, mc): 
            # reading cell value from source excel file 
            c = ws_src.cell(row = row_offset + i, column = col_offset + j) 
    
            # writing the read value to destination excel file 
            ws_dest.cell(row = target_row + i  , column = target_col + j).value = c.value 

        row_no = target_row + i
        col_no= target_col + j
        # 여기서 엑셀함수 컬럼 추가
        if i == 1:
            ws_dest.cell(row = row_no, column=col_no + 1).value = "증가비1"
            ws_dest.cell(row = row_no, column=col_no + 2).value = "증가비2"
            ws_dest.cell(row = row_no, column=col_no + 3).value = "증가비전체"
        if i > 2:
            col_no += 1
            prev_col = f"{get_column_letter(col_no-1)}{row_no}"
            pprev_col = f"{get_column_letter(col_no-2)}{row_no}"
            ppprev_col = f"{get_column_letter(col_no-3)}{row_no}"
            ws_dest.cell(row = row_no, column=col_no).value = f'=IFERROR({pprev_col}/{ppprev_col}*100,"N/A")'
            col_no += 1
            ws_dest.cell(row = row_no, column=col_no).value = f'=IFERROR({prev_col}/{pprev_col}*100,"N/A")'
            col_no += 1
            ws_dest.cell(row = row_no, column=col_no).value = f'=IFERROR(IF(OR({pprev_col}=0, AND({ppprev_col}=0, {prev_col}=0)), "N/A", IF({ppprev_col}=0, IF({prev_col}=0, 0, ({prev_col}-{pprev_col})/{pprev_col}), IF(AND({prev_col}=0, {pprev_col}=0), "N/A", IF({prev_col}=0, (({pprev_col}-{ppprev_col})/{ppprev_col}+(0-{pprev_col}/{pprev_col}))/2, (IFERROR(({pprev_col}-{ppprev_col})/{ppprev_col}, 0)+IFERROR(({prev_col}-{pprev_col})/{pprev_col}, 0))/2))))*100+100,"N/A")'
    
    target_row = target_row + mr
    target_col = target_col + mc + 3
    return target_row, target_col