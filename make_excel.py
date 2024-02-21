import mysql.connector, base64, openpyxl, io, os, re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, DEFAULT_FONT
from openpyxl.utils import get_column_letter
from appendSheet import appendSheet

def getExcelStreamByCompanyName(corp):
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password='admin',
        database="dart"  # Name of the database
    )
    
    cursor = mydb.cursor()
    # query = "SELECT  excel_1, excel_2, excel_3, stock_name FROM dart.corp_table where excel_3 is not null limit 5;"
    # query = f"SELECT  excel_1, excel_2, excel_3, stock_name FROM dart.corp_table where stock_name like '%{corp}%' limit 5;"
    query = f"SELECT excel_1, excel_2, excel_3, stock_name, bizr_no FROM dart.corp_table where bizr_no in "+corp+" limit 30;"

    cursor.execute(query)

    data = cursor.fetchall()

    # data[0][0] # => excel_1 (재무상태표)
    # data[0][1] # => excel_2 (손익계산서)
    # data[0][2] # => excel_3 (재무분석표)

    wb_dest = Workbook() # 새 워크북 생성
    ws_dest = wb_dest.active # 현재 활성화된 sheet 가져옴
    ws_dest.title = "원본" # sheet 의 이름을 변경

    start_row = 1
    start_col = 1
    row_offset = 1
    col_offset = 2
    # return str(len(data))
    for company in range(len(data)):
        ws_dest.cell(row = start_row, column = start_col, value = data[company][3] )
        start_row += 1
        if company > 0: 
            col_offset = 3
        for i in range(3):
            image = data[company][i]
 
            if image == None:
                # print(data[company][3])
                continue
        # Decode the string
            binary_data = base64.b64decode(image)
        
        # # 엑셀 파일 저장할 때
        # f = open("test.xlsx", 'wb')
        # f.write(binary_data)
        # f.close

        # OpenPyxl로 열 때
            workbook_xml = io.BytesIO(binary_data)
            workbook_xml.seek(0)
            wb = openpyxl.load_workbook(workbook_xml)
            ws_src = wb.active
            end_row, end_col = appendSheet(ws_dest, start_row, start_col, ws_src, row_offset, col_offset)
            start_row = end_row
        end_col += 1
        start_col = end_col
        start_row = 1

    #wb_dest.save("merged.xlsx")
    output = io.BytesIO()
    wb_dest.save(output)
    wb_dest.close()
    return output


def getExcelStreamByJobno(jobno):
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password='admin',
        database="dart"  # Name of the database
    )
    
    cursor = mydb.cursor()
    # query = "SELECT  excel_1, excel_2, excel_3, stock_name FROM dart.corp_table where excel_3 is not null limit 5;"
    # query = f"SELECT  excel_1, excel_2, excel_3, stock_name FROM dart.corp_table where stock_name like '%{corp}%' limit 5;"
    query = f"SELECT excel_1, excel_2, excel_3, stock_name, bizr_no FROM dart.corp_table a, dart.job_item b where b.jobno = '{jobno}' and a.bizr_no = b.biz_no;"

    cursor.execute(query)

    data = cursor.fetchall()

    # data[0][0] # => excel_1 (재무상태표)
    # data[0][1] # => excel_2 (손익계산서)
    # data[0][2] # => excel_3 (재무분석표)

    wb_dest = Workbook() # 새 워크북 생성
    _font = Font(name="Arial", sz=10, b=True)
    _font = Font(name="Arial", sz=9)
    {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    ws_dest = wb_dest.active # 현재 활성화된 sheet 가져옴
    ws_dest.title = "원본" # sheet 의 이름을 변경

    start_row = 1
    start_col = 1
    row_offset = 1
    col_offset = 2
    no_col = 4
    # return str(len(data))
    for company in range(len(data)):
        ws_dest.cell(row = start_row, column = start_col, value = data[company][3] )
        start_row += 1
        if company > 0: #두번째 회사부터
            col_offset = 3
            no_col = 3
        for i in range(3):
            image = data[company][i]
 
            if image == None:
                # print(data[company][3])
                continue
        # Decode the string
            binary_data = base64.b64decode(image)
        
        # # 엑셀 파일 저장할 때
        # f = open("test.xlsx", 'wb')
        # f.write(binary_data)
        # f.close

        # OpenPyxl로 열 때
            workbook_xml = io.BytesIO(binary_data)
            workbook_xml.seek(0)
            wb = openpyxl.load_workbook(workbook_xml)
            ws_src = wb.active
            end_row, end_col = appendSheet(ws_dest, start_row, start_col, ws_src, row_offset, col_offset, no_col)
            start_row = end_row
        start_col = end_col
        start_row = 1

    #wb_dest.save("merged.xlsx")

    output = io.BytesIO()
    wb_dest.save(output)
    wb_dest.close()
    ## ----------------엑셀 꾸미는 부분--------------
    wb = load_workbook(output) # sample.xlsx 파일에서 wb 을 불러옴
    ws = wb.active

    row_3 = ws[3]
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # ---------------------3행 굵게/가운데 정렬
    for i in range(2,len(row_3)+1):
        a1 = ws[f"{get_column_letter(i)}3"]
        a1.font = Font(bold=True)
        a1.alignment = Alignment(horizontal="center", vertical="center")
        a1.fill = PatternFill(fgColor="E7E5E6", fill_type="solid")
        a1.border = thin_border
        
    ## 각 칼럼에 대해서 모든 셀값의 문자열 개수에서 1.1만큼 곱한 것들 중 최대값을 계산한다.
    #for column_cells in ws.columns:
    #    length = max(len(str(cell.value))*1.1 for cell in column_cells)    
    #    ws.column_dimensions[column_cells[0].column_letter].width = length   

    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.1

    # 틀 고정
    ws.freeze_panes = "B4" 

    wb.save(output)
    return output