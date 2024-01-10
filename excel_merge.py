from openpyxl import Workbook, load_workbook
import re, os
wb = Workbook()
ws = wb.active
ws.title = "동종업종"
data = []
# cnt = 1

def same_area(cnt):
    # 현재 폴더에서 ETFI로 시작하는 파일 중 첫 번째 파일을 찾음
    filename = next((file for file in os.listdir('.') if re.match('^ETFI', file)), None)

    #파일 갯수 만큼 -> 한 업체당 엑셀 표 3개
    b = load_workbook(filename)
    bs = b["Sheet1"] # Dict 로 sheet 접근

    j=2 # 2행 부터
    while True:
        data.append( # 데이터 담는 과정
            [
                bs["B{}".format(j)].value, 
                bs["C{}".format(j)].value, 
                bs["D{}".format(j)].value, 
                bs["E{}".format(j)].value
            ]
        )
        j = j+1
        if bs["B{}".format(j)].value == None:
            break
    if cnt == 1:       
        for i in range(len(data)): # 데이터 쓰는 과정 (재무상태표)
            wb["동종업종"]["B{}".format(i+3)] = data[i][0]
            wb["동종업종"]["C{}".format(i+3)] = data[i][1]
            wb["동종업종"]["D{}".format(i+3)] = data[i][2]
            wb["동종업종"]["E{}".format(i+3)] = data[i][3]
        
    if cnt == 2:
        for i in range(len(data)): # 데이터 쓰는 과정 (손익계산서)
            wb["동종업종"]["B{}".format(i+803)] = data[i][0]
            wb["동종업종"]["C{}".format(i+803)] = data[i][1]
            wb["동종업종"]["D{}".format(i+803)] = data[i][2]
            wb["동종업종"]["E{}".format(i+803)] = data[i][3]
        
    if cnt == 3:
        for i in range(len(data)): # 데이터 쓰는 과정 (재무분석표)
            wb["동종업종"]["B{}".format(i+1146)] = data[i][0]
            wb["동종업종"]["C{}".format(i+1146)] = data[i][1]
            wb["동종업종"]["D{}".format(i+1146)] = data[i][2]
            wb["동종업종"]["E{}".format(i+1146)] = data[i][3]
        
    file = next((file for file in os.listdir('.') if re.match('^ETFI', file)), None) # 동종업종
    os.remove(file) # 추가한 파일 삭제

wb.save("동종업종.xlsx")
wb.close()