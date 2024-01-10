# import os
# os.system('pip install --upgrade selenium') selenium 항상 최신버전 유지
# pip install lxml
# pip install requests
import os, time, csv, re
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.chrome.options import Options

options = Options()
options.add_experimental_option('detach', True) # 브라우저 바로 닫힘 방지 옵션
# options.add_experimental_option('prefs', {'download.defalut_directory':r'C:\Users\rlaal\OneDrive\바탕 화면\Python_Rpa'}) # download 경로지정
browser = webdriver.Chrome(options=options)
com_nums = [] # 사업자 등록번호 list

pre_page = browser.find_element(By.XPATH, '//*[@id="psWrap"]/div[2]/ul/li[1]/a')
soup = BeautifulSoup(browser.page_source, 'html.parser')
def scrap():    # 동종업종 사업자등록번호 저장 반복문
    while True:
        try:
            # 페이지 로딩 대기
            wait = WebDriverWait(browser, 10)
            name_xpath = '//*[@id="tbody"]/tr[{}]/td[2]/span/a'.format(i) # 업체명 xpath
            elements = wait.until(EC.presence_of_all_elements_located((By.XPATH, name_xpath)))
            page_xpath = '//*[@id="psWrap"]/div[2]/ul/li[{}]/a'.format(i+1) # 다음 페이지 이동 xpath

            # 요소가 100개 미만일 경우, 끝 요소까지 반복 후 탈출
            for i in range(len(elements)):
                elem = elements[i]
                elem.click()
                corp = soup.select_one('#winCorpInfo > div.layerPop.layerPopM > div.cont > table > tbody > tr:nth-child(8) > td')
                com_num = corp.get_text()

                # com_nums 리스트에 값이 이미 있는 경우 continue
                if com_num in com_nums:
                    continue
                else:
                    com_nums.append(com_num)

                browser.find_element(By.LINK_TEXT, "닫기").click()

            # 다음 페이지 이동
            browser.find_element(By.XPATH, page_xpath).click()

        except Exception as e:
            print('페이지 끝')
            break
                
def same_area(cnt):
    # 현재 폴더에서 ETFI로 시작하는 파일 중 첫 번째 파일을 찾음
    filename = next((file for file in os.listdir('.') if re.match('^ETFI', file)), None)

    #파일 갯수 만큼 -> 한 업체당 엑셀 표 3개
    b = load_workbook(filename)
    bs = b["Sheet1"] # Dict 로 sheet 접근

    j=2 # 2행 부터
    while True:
        data.append( # 데이터 담는 과정
            [
                bs["B{}".format(j)].value, 
                bs["C{}".format(j)].value, 
                bs["D{}".format(j)].value, 
                bs["E{}".format(j)].value
            ]
        )
        j = j+1
        if bs["B{}".format(j)].value == None:
            break
    if cnt%3 == 1:       
        for i in range(len(data)): # 데이터 쓰는 과정 (재무상태표)
            wb["동종업종"]["B{}".format(i+3)] = data[i][0]
            wb["동종업종"]["C{}".format(i+3)] = data[i][1]
            wb["동종업종"]["D{}".format(i+3)] = data[i][2]
            wb["동종업종"]["E{}".format(i+3)] = data[i][3]
        cnt +=1
    if cnt%3 == 2:
        for i in range(len(data)): # 데이터 쓰는 과정 (손익계산서)
            wb["동종업종"]["B{}".format(i+803)] = data[i][0]
            wb["동종업종"]["C{}".format(i+803)] = data[i][1]
            wb["동종업종"]["D{}".format(i+803)] = data[i][2]
            wb["동종업종"]["E{}".format(i+803)] = data[i][3]
        cnt +=1
    if cnt%3 == 0:
        for i in range(len(data)): # 데이터 쓰는 과정 (재무분석표)
            wb["동종업종"]["B{}".format(i+1146)] = data[i][0]
            wb["동종업종"]["C{}".format(i+1146)] = data[i][1]
            wb["동종업종"]["D{}".format(i+1146)] = data[i][2]
            wb["동종업종"]["E{}".format(i+1146)] = data[i][3]
        cnt +=1

# browser.maximize_window()
wait = WebDriverWait(browser, 10)
url = "https://dart.fss.or.kr/dsab007/main.do?option=corp"

browser.get(url) # DART 통합검색 이동

browser.find_element(By.ID, "btnPlus").click() # 상세조건열기
browser.find_element(By.ID, "businessNm").click() # 업종 클릭
time.sleep(10) # 업종 선택 대기

browser.find_element(By.XPATH, '//*[@id="maxResultsCb"]/option[4]').click() # 조회건수 100 
browser.find_element(By.XPATH, '//*[@id="corporationType"]/option[2]').click() # 유가증권시장
browser.find_element(By.XPATH, '//*[@id="searchForm"]/div[2]/div[2]/a[1]').click() # 검색
time.sleep(2)
scrap() # 유가증권시장 scrap

browser.find_element(By.XPATH, '//*[@id="corporationType"]/option[3]').click() # 코스닥시장
browser.find_element(By.XPATH, '//*[@id="searchForm"]/div[2]/div[2]/a[1]').click() # 검색
scrap() # 코스닥시장 scrap

# ----------------------------------------------
url = "https://new.cretop.com/?h=1702989839177"
browser.get(url) # 크레탑으로 이동
time.sleep(2)

browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div[2]/div/div/div[2]/div[2]/div[2]/button').click() #닫기 버튼 
browser.find_element(By.ID, "idModel").send_keys("KAC100") # ID 입력
browser.find_element(By.ID, "pwModel").send_keys("kac100!!") # PW 입력
browser.find_element(By.CSS_SELECTOR, "[title='로그인']").click() # 로그인 

time.sleep(1)
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div[2]/div/div/div[2]/div/div/div[1]/div[2]/div/button[2]').click() # 본인인증 

time.sleep(1)
browser.find_element(By.LINK_TEXT, "휴대폰 인증").click() 

browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div/div[1]/ul/li[3]/span/label').click() #통신사 선택 # li[1] : SK li[2] : KT li[3] : LG
browser.find_element(By.ID, "PLCM050P1_agency-check-all").click() # 약관 전체동의
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[2]/div[2]/div/div/div[2]/div/div[2]/div/button').click() # 다음 버튼
browser.find_element(By.NAME, "PLCM050P2_username").send_keys("김민성")
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/div/div[1]/div/div[1]/div[2]/div/span[1]/label').click() #내국인
browser.find_element(By.ID, 'PLCM050P2_birth').send_keys("20010426")
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/div/div[1]/div/div[2]/div[2]/div/span[1]/label').click() # 성별(남자)선택
browser.find_element(By.ID, "PLCM050P2_phone2").send_keys("4373")
browser.find_element(By.ID, "PLCM050P2_phone3").send_keys("6549")
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[2]/div[2]/div/div/div[2]/div/div/div[1]/div/div/div[1]/div/div[3]/div[2]/div/button').click() #인증번호 요청
time.sleep(1)
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[3]/div[2]/div/div/ul/button').click() #인증번호가 전송되었습니다 [확인] 버튼클릭

# browser.find_element(By.ID, "PLCM050P2_certifyNumber").send_keys("인증번호6자리")
# -----------------------수동입력-----------------------------------------------------
'''
time.sleep(30) #문자 입력하는 시간

browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[2]/div[2]/div/div/div[2]/div/div/div[2]/div/button').click() #인증 버튼
time.sleep(1)
'''
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[3]/div[2]/div/div/ul/button').click() # 본인인증에 성공하였습니다 [확인] 버튼
time.sleep(1)

browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div[2]/div[2]/div/div/ul/button').click() # 다시 로그인 해주세요 [확인] 버튼
time.sleep(1)

browser.find_element(By.ID, "pwModel").send_keys("kac100!!") # PW 입력
browser.find_element(By.CSS_SELECTOR, "[title='로그인']").click() #로그인 클릭
time.sleep(2)

browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div[2]/div/div/ul/button').click() # 김민성님 로그인 되었습니다 [확인] 버튼
time.sleep(5)

''' 자동 로그아웃 
browser.find_element(By.XPATH,'//*[@id="app"]/div[2]/div/div[2]/div/div/ul/button').click() 자동로그아웃 처리 되었습니다 [확인]
browser.find_element(By.ID, "pwModel").send_keys("kac100!!") # PW 입력
browser.find_element(By.CSS_SELECTOR, "[title='로그인']").click() #로그인 클릭
time.sleep(2)
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div[2]/div/div/ul/button').click() # 김민성님 로그인 되었습니다 [확인] 버튼
'''
wb = Workbook()
ws = wb.active
ws.title = "동종업종"
data = []
cnt = 1


# -----xlsx 반복문 지점-----
for i in range(1, len(com_nums) + 1):
    if 
    browser.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/input[1]').send_keys(com_nums[i]) # 메인 화면 에서 첫 업체명 입력
    browser.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/button').click() # 검색버튼 클릭.
    browser.find_element(By.XPATH, '//*[@id="et-area"]/div/div[2]/ul/li/div/ul[3]/li[4]/a').click() # 재무 
    browser.find_element(By.XPATH, '//*[@id="etfi110m1"]/div/div[3]/div/div/div/div[1]/div/div[1]/div/div[1]/div[2]/div/span[2]/label').click() # 일반기업회계 
    browser.find_element(By.XPATH, '//*[@id="forms"]/option[2]').click() # 전계정 
    
    # browser.find_element(By.XPATH, '//*[@id="range"]/option[1]').click() # option[1] : 3년 option[2] : 5년
    
    browser.find_element(By.XPATH, '//*[@id="etfi110m1"]/div/div[3]/div/div/div/div[1]/div/div[1]/div/div[7]/button').click() # 조회하기
    time.sleep(1)

    same_area() # 재무제표 다운
    
    browser.find_element(By.LINK_TEXT, "손익계산서").click() 
    # browser.find_element(By.XPATH, '//*[@id="etfi110m1"]/div/div[3]/div/div/div/div[1]/div/div[2]/div/div/div/div[2]/div/div/div[2]/div/div[1]/div/div/button').click() 손익계산서 excel 다운로드 클릭

    same_area() # 손익계산서 다운

    browser.find_element(By.LINK_TEXT, "재무분석").click() 
    browser.find_element(By.XPATH, '//*[@id="etfi110m1"]/div/div[3]/div/div/div/div[4]/div/div[1]/div/div[2]/div[2]/div/span[2]/label').click() # 일반기업회계
    browser.find_element(By.XPATH, '//*[@id="etfi110m1"]/div/div[3]/div/div/div/div[4]/div/div[1]/div/div[2]/div[2]/div/span[2]/label').click() # 조회하기

    # browser.find_element(By.XPATH, '//*[@id="etfi110m1"]/div/div[3]/div/div/div/div[4]/div/div[2]/div[2]/div[1]/div/div/button').click() 재무분석 excel 다운로드 클릭

    same_area() # 재무분석 다운

    browser.find_element(By.XPATH, '//*[@id="app"]/div[1]/header/div/div[1]/div[1]/div[1]/input[1]').send_keys(com_nums[i])