import os
path=r'C:\Users\rlaal\OneDrive\바탕 화면\Python_Rpa\\'
def XlsxDelete():
    file_list = os.listdir(path) 
    for file in file_list: # 엑셀 파일들 삭제
        name, ext = os.path.splitext(file)
        if ext == '.xlsx':
            file_path = os.path.join(path, file)
            os.remove(file_path)

import os, time, csv, re, json, sys, pandas, base64
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
import mysql.connector
from selenium.common import exceptions
from selenium.common.exceptions import WebDriverException
from openpyxl import Workbook,load_workbook

def CretopExcel(corp):
    wb = Workbook() # 새 워크북 생성
    ws = wb.active # 현재 활성화된 sheet 가져옴


    def clearPerformanceLog():
        browser.get_log('performance')

    mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="admin",
    database="dart"
    )

    # mycursor = mydb.cursor()
    # mycursor.execute(same_area_query_sql)
    # companies = mycursor.fetchall()

    caps = DesiredCapabilities.CHROME.copy() 
    caps['goog:loggingPrefs'] = {"performance": "ALL"}
    options = ChromeOptions()
    options.add_experimental_option('debuggerAddress', '127.0.0.1:9222')
    browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options, desired_capabilities=caps)

    # -------------------------------------------------------
    url = "https://new.cretop.com"
    browser.get(url) # 크레탑으로 이동
    wait = WebDriverWait(browser, 10)
    time.sleep(3)


    for company in len(corp): 
        # print(company[2])  #corp_name
        browser.get("https://new.cretop.com/ET/SS/ETSS070M1") # 크레탑(기업)으로 이동
        time.sleep(4)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#app > div.wrap > header > div > div.top > div.search-form__area.sub > div.search-form > input[type=text]:nth-child(1)")))
        elem.send_keys(company) # 사업자 등록번호 입력
        time.sleep(4)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "i[class='search icon50 icon50 search']")))
        elem.click() # 검색 아이콘 클릭
        time.sleep(4)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#app > div.wrap > div.container > div > div > div:nth-child(2) > div:nth-child(3) > div.search-result__area > div > ul > li > div > ul.btn__list > li:nth-child(4) > a"))) # 재무 탭
        browser.execute_script("arguments[0].click()", elem) #  검색결과 페이지 로딩완료 및 재무단추 클릭
        # a[title='재무 페이지로 이동하기']
        time.sleep(4)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "label[for='account-standard-2']")))
        elem.click()     # 일반기업회계 단추 클릭
        time.sleep(3)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#forms > option:nth-child(2)")))
        elem.click() # 전계정
        time.sleep(3)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#etfi110m1 > div > div:nth-child(3) > div > div > div > div.tab-item.teb-item1 > div > div.inner__area.search-top__area > div > div.btn-wrap > button")))
        elem.click() # 조회하기
        time.sleep(4)

        file = next((file for file in os.listdir('.') if re.match('^ETFI', file)), None)
        if file:
            os.remove(file) # 추가한 파일 삭제
        clearPerformanceLog()
        time.sleep(3)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "button[class='btn excel-download']")))
        # button[class='btn excel-download']
        # browser.execute_script("arguments[0].click()", elem)
        elem.click() # 엑셀(재무상태표) 다운
        time.sleep(3) # 다운로드 대기

        filename = next((file for file in os.listdir('.') if re.match('^ETFI', file)), None)

        #파일 갯수 만큼 -> 한 업체당 엑셀 표 3개 반복
        wb_src = load_workbook(filename)

        # fp = open('ETFI112E1.xlsx','rb').read() # 재무상태표 insert
        # fp = base64.b64encode(fp)
        # args = (fp, company[0])
        # mycursor.execute(excel_insertion_sql,args)
        # mydb.commit()

        elem = wait.until(EC.presence_of_element_located((By.LINK_TEXT, "손익계산서")))
        elem.click()
        time.sleep(3)

        file = next((file for file in os.listdir('.') if re.match('^ETFI', file)), None)
        if file:
            os.remove(file) # 추가한 파일 삭제
        clearPerformanceLog()

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#etfi110m1 > div > div:nth-child(3) > div > div > div > div.tab-item.teb-item1 > div > div:nth-child(3) > div > div > div > div.tab-item.teb-item2 > div > div > div.table__area > div > div.content__head > div > div > button")))
        elem.click() # 손익계산 다운
        time.sleep(3) # 다운로드 대기

        # fp = open('ETFI112E1.xlsx','rb').read() # 손익계산서 insert
        # fp = base64.b64encode(fp)
        # args = (fp, company[0])
        # mycursor.execute(excel_2_sql,args) 
        # mydb.commit()

        elem = wait.until(EC.presence_of_element_located((By.LINK_TEXT, "재무분석")))
        elem.click()
        time.sleep(4)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "label[for='account-standard-4']")))
        elem.click() # 일반기업회계
        time.sleep(3)

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "button[class='btn default w-240 h-56']")))
        browser.execute_script("arguments[0].click()", elem) # 조회하기(재무분석)
        time.sleep(3)

        file = next((file for file in os.listdir('.') if re.match('^ETFI', file)), None)
        if file:
            os.remove(file) # 추가한 파일 삭제
        clearPerformanceLog()

        elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#etfi110m1 > div > div:nth-child(3) > div > div > div > div.tab-item.teb-item4 > div > div:nth-child(3) > div.table__area > div.content__head > div > div > button"))) # 재무분석 
        browser.execute_script("arguments[0].click()", elem) # 다운로드 클릭
        time.sleep(3) # 다운로드 대기

    # try:
    #     fp = open('ETFI114E1.xlsx','rb').read() # 재무분석 insert
    #     fp = base64.b64encode(fp)
    #     args = (fp, company[0])
    #     mycursor.execute(excel_3_sql,args) 
    #     mydb.commit()
    # except:
    #     pass
    # return wb.save("동종업종.xlsx") 